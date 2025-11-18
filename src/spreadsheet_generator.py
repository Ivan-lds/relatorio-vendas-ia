# Módulo gerador de planilhas Excel para venda
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# Cores padrão
COR_PRIMARIA = "636EFA"
COR_SECUNDARIA = "4E5FD3"
COR_SUCESSO = "28A745"
COR_ALERTA = "FFC107"
COR_ERRO = "DC3545"
COR_FUNDO = "F8F9FA"
COR_TEXTO = "212529"

def criar_estilos():
    """Define estilos padrão para as planilhas"""
    styles = {
        'titulo': Font(name='Calibri', size=16, bold=True, color="FFFFFF"),
        'subtitulo': Font(name='Calibri', size=14, bold=True, color=COR_TEXTO),
        'cabecalho': Font(name='Calibri', size=11, bold=True, color="FFFFFF"),
        'texto': Font(name='Calibri', size=11, color=COR_TEXTO),
        'numero': Font(name='Calibri', size=11, color=COR_TEXTO),
        'total': Font(name='Calibri', size=11, bold=True, color=COR_TEXTO),
        'fundo_titulo': PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA, fill_type='solid'),
        'fundo_cabecalho': PatternFill(start_color=COR_SECUNDARIA, end_color=COR_SECUNDARIA, fill_type='solid'),
        'fundo_total': PatternFill(start_color=COR_FUNDO, end_color=COR_FUNDO, fill_type='solid'),
        'borda': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'centralizar': Alignment(horizontal='center', vertical='center'),
        'direita': Alignment(horizontal='right', vertical='center'),
        'esquerda': Alignment(horizontal='left', vertical='center')
    }
    return styles

def aplicar_bordas(ws, min_row, max_row, min_col, max_col):
    """Aplica bordas a um intervalo de células"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border

def criar_controle_financeiro_completo():
    """Cria planilha de Controle Financeiro Completo"""
    wb = Workbook()
    styles = criar_estilos()
    
    # Remover sheet padrão
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    # Título
    ws_dash['A1'] = "CONTROLE FINANCEIRO COMPLETO"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:H1')
    
    # Resumo Financeiro
    ws_dash['A3'] = "RESUMO FINANCEIRO"
    ws_dash['A3'].font = styles['subtitulo']
    ws_dash['A3'].fill = styles['fundo_total']
    
    resumo_labels = ['Receitas do Mês', 'Despesas do Mês', 'Saldo Atual', 'Saldo Previsto']
    resumo_cols = ['A', 'C', 'E', 'G']
    resumo_val_cols = ['B', 'D', 'F', 'H']
    
    for i, (label, col, val_col) in enumerate(zip(resumo_labels, resumo_cols, resumo_val_cols)):
        row = 5 + i
        ws_dash[f'{col}{row}'] = label
        ws_dash[f'{col}{row}'].font = styles['texto']
        ws_dash[f'{val_col}{row}'] = f"=SUMIF(FluxoCaixa!E:E,\">0\",FluxoCaixa!F:F)" if i == 0 else \
                                     f"=SUMIF(FluxoCaixa!E:E,\"<0\",FluxoCaixa!F:F)" if i == 1 else \
                                     f"=SUM(FluxoCaixa!F:F)" if i == 2 else \
                                     f"={val_col}5+{val_col}6"
        ws_dash[f'{val_col}{row}'].number_format = 'R$ #,##0.00'
        ws_dash[f'{val_col}{row}'].font = styles['numero']
    
    # ============ SHEET: FLUXO DE CAIXA ============
    ws_fluxo = wb.create_sheet("FluxoCaixa")
    
    # Cabeçalho
    headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Saldo Anterior', 'Valor', 'Saldo Atual']
    for col, header in enumerate(headers, 1):
        cell = ws_fluxo.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    # Formato de colunas
    ws_fluxo.column_dimensions['A'].width = 12
    ws_fluxo.column_dimensions['B'].width = 30
    ws_fluxo.column_dimensions['C'].width = 20
    ws_fluxo.column_dimensions['D'].width = 12
    ws_fluxo.column_dimensions['E'].width = 15
    ws_fluxo.column_dimensions['F'].width = 15
    ws_fluxo.column_dimensions['G'].width = 15
    
    # Fórmulas de exemplo
    # Linha 2 (primeira entrada)
    ws_fluxo['A2'] = datetime.now().strftime('%d/%m/%Y')
    ws_fluxo['B2'] = "Saldo Inicial"
    ws_fluxo['C2'] = "Inicial"
    ws_fluxo['D2'] = "Entrada"
    ws_fluxo['E2'] = 0
    ws_fluxo['F2'] = 10000  # Valor inicial
    ws_fluxo['G2'] = "=E2+F2"
    
    # Formatação
    ws_fluxo['A2'].number_format = 'dd/mm/yyyy'
    ws_fluxo['E2'].number_format = 'R$ #,##0.00'
    ws_fluxo['F2'].number_format = 'R$ #,##0.00'
    ws_fluxo['G2'].number_format = 'R$ #,##0.00'
    
    # Linha 3 (exemplo)
    ws_fluxo['A3'] = (datetime.now() + timedelta(days=1)).strftime('%d/%m/%Y')
    ws_fluxo['B3'] = "Venda de Produtos"
    ws_fluxo['C3'] = "Receita"
    ws_fluxo['D3'] = "Entrada"
    ws_fluxo['E3'] = "=G2"
    ws_fluxo['F3'] = 2500
    ws_fluxo['G3'] = "=E3+F3"
    
    ws_fluxo['A3'].number_format = 'dd/mm/yyyy'
    ws_fluxo['E3'].number_format = 'R$ #,##0.00'
    ws_fluxo['F3'].number_format = 'R$ #,##0.00'
    ws_fluxo['G3'].number_format = 'R$ #,##0.00'
    
    aplicar_bordas(ws_fluxo, 1, 100, 1, 7)
    
    # ============ SHEET: CONTAS A PAGAR ============
    ws_pagar = wb.create_sheet("Contas a Pagar")
    
    headers_pagar = ['Vencimento', 'Descrição', 'Categoria', 'Valor', 'Status', 'Data Pagamento']
    for col, header in enumerate(headers_pagar, 1):
        cell = ws_pagar.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    ws_pagar.column_dimensions['A'].width = 12
    ws_pagar.column_dimensions['B'].width = 30
    ws_pagar.column_dimensions['C'].width = 20
    ws_pagar.column_dimensions['D'].width = 15
    ws_pagar.column_dimensions['E'].width = 12
    ws_pagar.column_dimensions['F'].width = 12
    
    # Total
    ws_pagar['C100'] = "TOTAL:"
    ws_pagar['C100'].font = styles['total']
    ws_pagar['D100'] = "=SUM(D2:D99)"
    ws_pagar['D100'].number_format = 'R$ #,##0.00'
    ws_pagar['D100'].font = styles['total']
    
    aplicar_bordas(ws_pagar, 1, 100, 1, 6)
    
    # ============ SHEET: CONTAS A RECEBER ============
    ws_receber = wb.create_sheet("Contas a Receber")
    
    headers_receber = ['Vencimento', 'Descrição', 'Categoria', 'Valor', 'Status', 'Data Recebimento']
    for col, header in enumerate(headers_receber, 1):
        cell = ws_receber.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    ws_receber.column_dimensions['A'].width = 12
    ws_receber.column_dimensions['B'].width = 30
    ws_receber.column_dimensions['C'].width = 20
    ws_receber.column_dimensions['D'].width = 15
    ws_receber.column_dimensions['E'].width = 12
    ws_receber.column_dimensions['F'].width = 12
    
    # Total
    ws_receber['C100'] = "TOTAL:"
    ws_receber['C100'].font = styles['total']
    ws_receber['D100'] = "=SUM(D2:D99)"
    ws_receber['D100'].number_format = 'R$ #,##0.00'
    ws_receber['D100'].font = styles['total']
    
    aplicar_bordas(ws_receber, 1, 100, 1, 6)
    
    return wb

def criar_controle_vendas_comissoes():
    """Cria planilha de Controle de Vendas e Comissões"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "CONTROLE DE VENDAS E COMISSÕES"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:F1')
    
    # Métricas
    metricas = [
        ('Total de Vendas', "=SUM(Vendas!F:F)"),
        ('Total de Comissões', "=SUM(Vendas!G:G)"),
        ('Média por Venda', "=A3/A5"),
        ('Total de Vendedores', "=COUNTA(Vendedores!B:B)-1"),
        ('Total de Vendas (Qtd)', "=COUNTA(Vendas!A:A)-1"),
        ('Ticket Médio', "=A3/A7")
    ]
    
    for i, (label, formula) in enumerate(metricas):
        row = 3 + i
        ws_dash[f'A{row}'] = label
        ws_dash[f'A{row}'].font = styles['texto']
        ws_dash[f'B{row}'] = formula
        ws_dash[f'B{row}'].number_format = 'R$ #,##0.00' if i < 3 or i == 5 else '#,##0'
        ws_dash[f'B{row}'].font = styles['numero']
    
    # ============ SHEET: VENDAS ============
    ws_vendas = wb.create_sheet("Vendas")
    
    headers = ['Data', 'Vendedor', 'Cliente', 'Produto', 'Quantidade', 'Valor Total', 'Comissão (%)', 'Valor Comissão']
    for col, header in enumerate(headers, 1):
        cell = ws_vendas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    # Largura das colunas
    for i, width in enumerate([12, 20, 25, 25, 12, 15, 12, 15], 1):
        ws_vendas.column_dimensions[get_column_letter(i)].width = width
    
    # Exemplo de fórmula de comissão
    # Coluna H (Valor Comissão) = F2 * G2 / 100
    
    aplicar_bordas(ws_vendas, 1, 500, 1, 8)
    
    # ============ SHEET: VENDEDORES ============
    ws_vendedores = wb.create_sheet("Vendedores")
    
    headers_vend = ['Código', 'Nome', 'CPF', 'Email', 'Telefone', 'Comissão (%)', 'Total Vendas', 'Total Comissões']
    for col, header in enumerate(headers_vend, 1):
        cell = ws_vendedores.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 25, 15, 25, 15, 12, 15, 15], 1):
        ws_vendedores.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_vendedores, 1, 100, 1, 8)
    
    # ============ SHEET: PRODUTOS ============
    ws_produtos = wb.create_sheet("Produtos")
    
    headers_prod = ['Código', 'Nome', 'Categoria', 'Preço Unitário', 'Estoque']
    for col, header in enumerate(headers_prod, 1):
        cell = ws_produtos.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 20, 15, 12], 1):
        ws_produtos.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_produtos, 1, 500, 1, 5)
    
    return wb

def criar_crm_basico():
    """Cria planilha de CRM Básico"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "CRM BÁSICO - GESTÃO DE CLIENTES"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:E1')
    
    # ============ SHEET: CLIENTES ============
    ws_clientes = wb.create_sheet("Clientes")
    
    headers = ['Código', 'Nome/Razão Social', 'CPF/CNPJ', 'Email', 'Telefone', 'Cidade', 'Estado', 
               'Segmento', 'Status', 'Data Cadastro', 'Último Contato', 'Valor Total Vendas']
    for col, header in enumerate(headers, 1):
        cell = ws_clientes.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 18, 25, 15, 20, 10, 15, 12, 12, 12, 15], 1):
        ws_clientes.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_clientes, 1, 1000, 1, 12)
    
    # ============ SHEET: PIPELINE ============
    ws_pipeline = wb.create_sheet("Pipeline")
    
    headers_pipe = ['Oportunidade', 'Cliente', 'Valor', 'Probabilidade', 'Etapa', 'Data Prevista', 
                    'Responsável', 'Observações']
    for col, header in enumerate(headers_pipe, 1):
        cell = ws_pipeline.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([30, 25, 15, 12, 15, 12, 20, 40], 1):
        ws_pipeline.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_pipeline, 1, 500, 1, 8)
    
    # ============ SHEET: INTERAÇÕES ============
    ws_inter = wb.create_sheet("Interações")
    
    headers_inter = ['Data', 'Cliente', 'Tipo', 'Assunto', 'Descrição', 'Responsável', 'Próximo Follow-up']
    for col, header in enumerate(headers_inter, 1):
        cell = ws_inter.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 15, 30, 40, 20, 12], 1):
        ws_inter.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_inter, 1, 1000, 1, 7)
    
    return wb

def criar_gestao_orcamento_familiar():
    """Cria planilha de Gestão de Orçamento Familiar"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "GESTÃO DE ORÇAMENTO FAMILIAR"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:D1')
    
    # Resumo
    resumo = [
        ('Renda Total', "=SUM(Receitas!C:C)"),
        ('Despesas Total', "=SUM(Despesas!D:D)"),
        ('Saldo', "=A3-A4"),
        ('Meta de Economia', 1000),
        ('Economia Real', "=A5-A7")
    ]
    
    for i, (label, formula) in enumerate(resumo):
        row = 3 + i
        ws_dash[f'A{row}'] = label
        ws_dash[f'A{row}'].font = styles['texto']
        ws_dash[f'B{row}'] = formula
        if isinstance(formula, (int, float)):
            ws_dash[f'B{row}'].value = formula
        ws_dash[f'B{row}'].number_format = 'R$ #,##0.00'
        ws_dash[f'B{row}'].font = styles['numero']
    
    # ============ SHEET: RECEITAS ============
    ws_receitas = wb.create_sheet("Receitas")
    
    headers = ['Data', 'Descrição', 'Categoria', 'Valor', 'Mês']
    for col, header in enumerate(headers, 1):
        cell = ws_receitas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 30, 20, 15, 10], 1):
        ws_receitas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_receitas, 1, 500, 1, 5)
    
    # ============ SHEET: DESPESAS ============
    ws_despesas = wb.create_sheet("Despesas")
    
    headers = ['Data', 'Descrição', 'Categoria', 'Valor', 'Mês']
    for col, header in enumerate(headers, 1):
        cell = ws_despesas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 30, 20, 15, 10], 1):
        ws_despesas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_despesas, 1, 500, 1, 5)
    
    # ============ SHEET: INVESTIMENTOS ============
    ws_invest = wb.create_sheet("Investimentos")
    
    headers_invest = ['Tipo', 'Instituição', 'Valor Investido', 'Rentabilidade (%)', 'Valor Atual', 'Data Aplicação']
    for col, header in enumerate(headers_invest, 1):
        cell = ws_invest.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([20, 25, 15, 12, 15, 12], 1):
        ws_invest.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_invest, 1, 100, 1, 6)
    
    return wb

def criar_calculadora_emprestimos():
    """Cria planilha de Calculadora de Empréstimos"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: SIMULADOR ============
    ws_sim = wb.create_sheet("Simulador")
    
    ws_sim['A1'] = "CALCULADORA DE EMPRÉSTIMOS E FINANCIAMENTOS"
    ws_sim['A1'].font = styles['titulo']
    ws_sim['A1'].fill = styles['fundo_titulo']
    ws_sim['A1'].alignment = styles['centralizar']
    ws_sim.merge_cells('A1:D1')
    
    # Dados de entrada
    ws_sim['A3'] = "DADOS DO EMPRÉSTIMO"
    ws_sim['A3'].font = styles['subtitulo']
    
    inputs = [
        ('Valor do Empréstimo (R$):', 'B4', 10000),
        ('Taxa de Juros (% ao mês):', 'B5', 1.5),
        ('Número de Parcelas:', 'B6', 12),
        ('Sistema:', 'B7', 'PRICE')
    ]
    
    for i, (label, cell_ref, default) in enumerate(inputs):
        row = 4 + i
        ws_sim[f'A{row}'] = label
        ws_sim[f'A{row}'].font = styles['texto']
        ws_sim[cell_ref] = default
        ws_sim[cell_ref].font = styles['numero']
        if i < 3:
            ws_sim[cell_ref].number_format = '#,##0.00' if i == 0 else '#,##0.00' if i == 1 else '#,##0'
    
    # Resultados
    ws_sim['A9'] = "RESULTADOS"
    ws_sim['A9'].font = styles['subtitulo']
    
    resultados = [
        ('Valor da Parcela:', 'B10', "=PGTO(B5/100,B6,-B4)"),
        ('Total Pago:', 'B11', "=B10*B6"),
        ('Total de Juros:', 'B12', "=B11-B4")
    ]
    
    for label, cell_ref, formula in resultados:
        ws_sim[cell_ref.replace('B', 'A')] = label
        ws_sim[cell_ref.replace('B', 'A')].font = styles['texto']
        ws_sim[cell_ref] = formula
        ws_sim[cell_ref].number_format = 'R$ #,##0.00'
        ws_sim[cell_ref].font = styles['numero']
    
    # ============ SHEET: TABELA PRICE ============
    ws_price = wb.create_sheet("Tabela Price")
    
    headers = ['Parcela', 'Saldo Devedor Inicial', 'Valor da Parcela', 'Juros', 'Amortização', 'Saldo Devedor Final']
    for col, header in enumerate(headers, 1):
        cell = ws_price.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 18, 18, 15, 15, 18], 1):
        ws_price.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_price, 1, 300, 1, 6)
    
    # ============ SHEET: TABELA SAC ============
    ws_sac = wb.create_sheet("Tabela SAC")
    
    for col, header in enumerate(headers, 1):
        cell = ws_sac.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 18, 18, 15, 15, 18], 1):
        ws_sac.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_sac, 1, 300, 1, 6)
    
    return wb

def criar_controle_ponto_folha():
    """Cria planilha de Controle de Ponto e Folha de Pagamento"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "CONTROLE DE PONTO E FOLHA DE PAGAMENTO"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:E1')
    
    # ============ SHEET: FUNCIONÁRIOS ============
    ws_func = wb.create_sheet("Funcionários")
    
    headers = ['Código', 'Nome', 'CPF', 'Cargo', 'Salário Base', 'Carga Horária', 'Data Admissão']
    for col, header in enumerate(headers, 1):
        cell = ws_func.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 15, 20, 15, 12, 12], 1):
        ws_func.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_func, 1, 200, 1, 7)
    
    # ============ SHEET: PONTO ============
    ws_ponto = wb.create_sheet("Ponto")
    
    headers_ponto = ['Data', 'Funcionário', 'Entrada', 'Saída', 'Entrada', 'Saída', 'Horas Trabalhadas', 'Horas Extras']
    for col, header in enumerate(headers_ponto, 1):
        cell = ws_ponto.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 10, 10, 10, 10, 15, 12], 1):
        ws_ponto.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_ponto, 1, 500, 1, 8)
    
    # ============ SHEET: FOLHA DE PAGAMENTO ============
    ws_folha = wb.create_sheet("Folha de Pagamento")
    
    headers_folha = ['Mês/Ano', 'Funcionário', 'Salário Base', 'Horas Extras', 'Valor Horas Extras', 
                     'INSS', 'IRRF', 'FGTS', 'Descontos', 'Valor Líquido']
    for col, header in enumerate(headers_folha, 1):
        cell = ws_folha.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 15, 12, 15, 12, 12, 12, 12, 15], 1):
        ws_folha.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_folha, 1, 500, 1, 10)
    
    return wb

def criar_gestao_funcionarios_beneficios():
    """Cria planilha de Gestão de Funcionários e Benefícios"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: FUNCIONÁRIOS ============
    ws_func = wb.create_sheet("Funcionários")
    
    headers = ['Código', 'Nome', 'CPF', 'Cargo', 'Departamento', 'Data Admissão', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_func.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 15, 20, 20, 12, 12], 1):
        ws_func.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_func, 1, 200, 1, 7)
    
    # ============ SHEET: FÉRIAS ============
    ws_ferias = wb.create_sheet("Férias")
    
    headers_ferias = ['Funcionário', 'Período Aquisitivo', 'Data Início', 'Data Fim', 'Dias', 'Status']
    for col, header in enumerate(headers_ferias, 1):
        cell = ws_ferias.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([25, 18, 12, 12, 10, 12], 1):
        ws_ferias.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_ferias, 1, 200, 1, 6)
    
    # ============ SHEET: BENEFÍCIOS ============
    ws_benef = wb.create_sheet("Benefícios")
    
    headers_benef = ['Funcionário', 'Tipo de Benefício', 'Valor', 'Data Início', 'Data Fim']
    for col, header in enumerate(headers_benef, 1):
        cell = ws_benef.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([25, 25, 15, 12, 12], 1):
        ws_benef.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_benef, 1, 200, 1, 5)
    
    return wb

def criar_controle_estoque_completo():
    """Cria planilha de Controle de Estoque Completo"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "CONTROLE DE ESTOQUE COMPLETO"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:D1')
    
    # ============ SHEET: PRODUTOS ============
    ws_prod = wb.create_sheet("Produtos")
    
    headers = ['Código', 'Nome', 'Categoria', 'Unidade', 'Estoque Mínimo', 'Estoque Atual', 'Valor Unitário', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws_prod.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 20, 10, 12, 12, 15, 15], 1):
        ws_prod.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_prod, 1, 1000, 1, 8)
    
    # ============ SHEET: ENTRADAS ============
    ws_ent = wb.create_sheet("Entradas")
    
    headers_ent = ['Data', 'Produto', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Fornecedor', 'NF']
    for col, header in enumerate(headers_ent, 1):
        cell = ws_ent.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 12, 15, 15, 25, 15], 1):
        ws_ent.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_ent, 1, 500, 1, 7)
    
    # ============ SHEET: SAÍDAS ============
    ws_sai = wb.create_sheet("Saídas")
    
    headers_sai = ['Data', 'Produto', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Destino']
    for col, header in enumerate(headers_sai, 1):
        cell = ws_sai.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 12, 15, 15, 25], 1):
        ws_sai.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_sai, 1, 500, 1, 6)
    
    return wb

def criar_gestao_fornecedores_compras():
    """Cria planilha de Gestão de Fornecedores e Compras"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: FORNECEDORES ============
    ws_forn = wb.create_sheet("Fornecedores")
    
    headers = ['Código', 'Nome/Razão Social', 'CNPJ', 'Email', 'Telefone', 'Cidade', 'Estado', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_forn.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 18, 25, 15, 20, 10, 12], 1):
        ws_forn.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_forn, 1, 200, 1, 8)
    
    # ============ SHEET: PEDIDOS DE COMPRA ============
    ws_ped = wb.create_sheet("Pedidos de Compra")
    
    headers_ped = ['Número', 'Data', 'Fornecedor', 'Produto', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Status']
    for col, header in enumerate(headers_ped, 1):
        cell = ws_ped.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 12, 25, 25, 12, 15, 15, 12], 1):
        ws_ped.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_ped, 1, 500, 1, 8)
    
    # ============ SHEET: COTAÇÕES ============
    ws_cot = wb.create_sheet("Cotações")
    
    headers_cot = ['Produto', 'Fornecedor 1', 'Valor 1', 'Fornecedor 2', 'Valor 2', 'Fornecedor 3', 'Valor 3', 'Melhor Opção']
    for col, header in enumerate(headers_cot, 1):
        cell = ws_cot.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([25, 20, 15, 20, 15, 20, 15, 20], 1):
        ws_cot.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_cot, 1, 200, 1, 8)
    
    return wb

def criar_gerenciamento_projetos():
    """Cria planilha de Gerenciamento de Projetos"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "GERENCIAMENTO DE PROJETOS"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:E1')
    
    # ============ SHEET: PROJETOS ============
    ws_proj = wb.create_sheet("Projetos")
    
    headers = ['Código', 'Nome', 'Cliente', 'Responsável', 'Data Início', 'Data Fim', 'Status', 'Progresso (%)']
    for col, header in enumerate(headers, 1):
        cell = ws_proj.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 25, 20, 12, 12, 12, 12], 1):
        ws_proj.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_proj, 1, 200, 1, 8)
    
    # ============ SHEET: TAREFAS ============
    ws_tarefas = wb.create_sheet("Tarefas")
    
    headers_tar = ['Projeto', 'Tarefa', 'Responsável', 'Data Início', 'Data Fim', 'Status', 'Progresso (%)', 'Observações']
    for col, header in enumerate(headers_tar, 1):
        cell = ws_tarefas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([20, 30, 20, 12, 12, 12, 12, 30], 1):
        ws_tarefas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_tarefas, 1, 500, 1, 8)
    
    # ============ SHEET: RECURSOS ============
    ws_rec = wb.create_sheet("Recursos")
    
    headers_rec = ['Projeto', 'Recurso', 'Quantidade', 'Custo Unitário', 'Custo Total']
    for col, header in enumerate(headers_rec, 1):
        cell = ws_rec.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([20, 25, 12, 15, 15], 1):
        ws_rec.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_rec, 1, 300, 1, 5)
    
    return wb

def criar_agenda_compromissos():
    """Cria planilha de Agenda de Compromissos e Tarefas"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: AGENDA ============
    ws_agenda = wb.create_sheet("Agenda")
    
    headers = ['Data', 'Hora', 'Compromisso', 'Local', 'Status', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws_agenda.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 10, 30, 20, 12, 40], 1):
        ws_agenda.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_agenda, 1, 500, 1, 6)
    
    # ============ SHEET: TAREFAS ============
    ws_tarefas = wb.create_sheet("Tarefas")
    
    headers_tar = ['Tarefa', 'Prioridade', 'Prazo', 'Status', 'Data Conclusão', 'Observações']
    for col, header in enumerate(headers_tar, 1):
        cell = ws_tarefas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([30, 12, 12, 12, 12, 40], 1):
        ws_tarefas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_tarefas, 1, 500, 1, 6)
    
    return wb

def criar_controle_estudos():
    """Cria planilha de Controle de Estudos e Aprendizado"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "CONTROLE DE ESTUDOS E APRENDIZADO"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:D1')
    
    # ============ SHEET: ESTUDOS ============
    ws_est = wb.create_sheet("Estudos")
    
    headers = ['Data', 'Matéria/Assunto', 'Tempo (horas)', 'Conteúdo Estudado', 'Dificuldade', 'Anotações']
    for col, header in enumerate(headers, 1):
        cell = ws_est.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 12, 40, 12, 40], 1):
        ws_est.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_est, 1, 500, 1, 6)
    
    # ============ SHEET: METAS ============
    ws_metas = wb.create_sheet("Metas")
    
    headers_metas = ['Meta', 'Data Início', 'Data Fim', 'Horas Planejadas', 'Horas Realizadas', 'Progresso (%)', 'Status']
    for col, header in enumerate(headers_metas, 1):
        cell = ws_metas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([30, 12, 12, 15, 15, 12, 12], 1):
        ws_metas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_metas, 1, 200, 1, 7)
    
    # ============ SHEET: CALENDÁRIO PROVAS ============
    ws_provas = wb.create_sheet("Calendário Provas")
    
    headers_provas = ['Data', 'Matéria', 'Tipo', 'Peso', 'Nota', 'Status']
    for col, header in enumerate(headers_provas, 1):
        cell = ws_provas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 15, 10, 10, 12], 1):
        ws_provas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_provas, 1, 200, 1, 6)
    
    return wb

def criar_analise_vendas_dashboard():
    """Cria planilha de Análise de Vendas com Dashboard"""
    wb = Workbook()
    styles = criar_estilos()
    
    wb.remove(wb.active)
    
    # ============ SHEET: DASHBOARD ============
    ws_dash = wb.create_sheet("Dashboard")
    
    ws_dash['A1'] = "ANÁLISE DE VENDAS - DASHBOARD"
    ws_dash['A1'].font = styles['titulo']
    ws_dash['A1'].fill = styles['fundo_titulo']
    ws_dash['A1'].alignment = styles['centralizar']
    ws_dash.merge_cells('A1:H1')
    
    # KPIs
    kpis = [
        ('Total de Vendas', "=SUM(Vendas!F:F)"),
        ('Ticket Médio', "=AVERAGE(Vendas!F:F)"),
        ('Total de Clientes', "=COUNTA(Clientes!B:B)-1"),
        ('Taxa de Conversão (%)', "=Vendas!F2/Clientes!B2*100"),
        ('Vendas do Mês', "=SUMIF(Vendas!A:A,\">=\"&DATA(ANO(HOJE()),MÊS(HOJE()),1),Vendas!F:F)"),
        ('Meta do Mês', 50000),
        ('% da Meta', "=A7/A8*100")
    ]
    
    for i, (label, formula) in enumerate(kpis):
        row = 3 + i
        ws_dash[f'A{row}'] = label
        ws_dash[f'A{row}'].font = styles['texto']
        ws_dash[f'B{row}'] = formula
        if isinstance(formula, (int, float)):
            ws_dash[f'B{row}'].value = formula
        ws_dash[f'B{row}'].number_format = 'R$ #,##0.00' if i < 2 or i == 4 or i == 5 else '#,##0.00' if i == 3 or i == 6 else '#,##0'
        ws_dash[f'B{row}'].font = styles['numero']
    
    # ============ SHEET: VENDAS ============
    ws_vendas = wb.create_sheet("Vendas")
    
    headers = ['Data', 'Cliente', 'Produto', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Vendedor', 'Canal']
    for col, header in enumerate(headers, 1):
        cell = ws_vendas.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([12, 25, 25, 12, 15, 15, 20, 15], 1):
        ws_vendas.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_vendas, 1, 1000, 1, 8)
    
    # ============ SHEET: CLIENTES ============
    ws_clientes = wb.create_sheet("Clientes")
    
    headers_cli = ['Código', 'Nome', 'Segmento', 'Total Compras', 'Última Compra', 'Status']
    for col, header in enumerate(headers_cli, 1):
        cell = ws_clientes.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 20, 15, 12, 12], 1):
        ws_clientes.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_clientes, 1, 500, 1, 6)
    
    # ============ SHEET: PRODUTOS ============
    ws_prod = wb.create_sheet("Produtos")
    
    headers_prod = ['Código', 'Nome', 'Categoria', 'Quantidade Vendida', 'Receita Total']
    for col, header in enumerate(headers_prod, 1):
        cell = ws_prod.cell(row=1, column=col)
        cell.value = header
        cell.font = styles['cabecalho']
        cell.fill = styles['fundo_cabecalho']
        cell.alignment = styles['centralizar']
    
    for i, width in enumerate([10, 30, 20, 15, 15], 1):
        ws_prod.column_dimensions[get_column_letter(i)].width = width
    
    aplicar_bordas(ws_prod, 1, 500, 1, 5)
    
    return wb

def gerar_todas_planilhas(output_dir="planilhas"):
    """Gera todas as planilhas do catálogo"""
    os.makedirs(output_dir, exist_ok=True)
    
    planilhas = {
        "Controle_Financeiro_Completo.xlsx": criar_controle_financeiro_completo,
        "Gestao_Orcamento_Familiar.xlsx": criar_gestao_orcamento_familiar,
        "Calculadora_Emprestimos_Financiamentos.xlsx": criar_calculadora_emprestimos,
        "Controle_Vendas_Comissoes.xlsx": criar_controle_vendas_comissoes,
        "CRM_Basico_Excel.xlsx": criar_crm_basico,
        "Analise_Vendas_Dashboard.xlsx": criar_analise_vendas_dashboard,
        "Controle_Ponto_Folha_Pagamento.xlsx": criar_controle_ponto_folha,
        "Gestao_Funcionarios_Beneficios.xlsx": criar_gestao_funcionarios_beneficios,
        "Controle_Estoque_Completo.xlsx": criar_controle_estoque_completo,
        "Gestao_Fornecedores_Compras.xlsx": criar_gestao_fornecedores_compras,
        "Gerenciamento_Projetos.xlsx": criar_gerenciamento_projetos,
        "Agenda_Compromissos_Tarefas.xlsx": criar_agenda_compromissos,
        "Controle_Estudos_Aprendizado.xlsx": criar_controle_estudos
    }
    
    resultados = []
    
    for nome_arquivo, funcao_geradora in planilhas.items():
        try:
            caminho = os.path.join(output_dir, nome_arquivo)
            wb = funcao_geradora()
            wb.save(caminho)
            resultados.append(f"✅ {nome_arquivo} criada com sucesso!")
        except Exception as e:
            resultados.append(f"❌ Erro ao criar {nome_arquivo}: {str(e)}")
    
    return resultados

if __name__ == "__main__":
    print("Gerando planilhas Excel...")
    resultados = gerar_todas_planilhas()
    for resultado in resultados:
        print(resultado)

